from colorsys import hls_to_rgb, rgb_to_hls

import matplotlib.pyplot as plt
import openpyxl
from openpyxl.utils import column_index_from_string

# Here Oliver, I added something to your code!

# https://bitbucket.org/openpyxl/openpyxl/issues/987/add-utility-functions-for-colors-to-help

RGBMAX = 0xFF  # Corresponds to 255
HLSMAX = 240  # MS excel's tint function expects that HLS is base 240. see:
# https://social.msdn.microsoft.com/Forums/en-US/e9d8c136-6d62-4098-9b1b-dac786149f43/excel-color-tint-algorithm-incorrect?forum=os_binaryfile#d3c2ac95-52e0-476b-86f1-e2a697f24969


def rgb_to_ms_hls(red, green=None, blue=None):
    """Converts rgb values in range (0,1) or a hex string of the form '[#aa]rrggbb' to HLSMAX based HLS, (alpha values are ignored)"""
    if green is None:
        if isinstance(red, str):
            if len(red) > 6:
                red = red[-6:]  # Ignore preceding '#' and alpha values
            blue = int(red[4:], 16) / RGBMAX
            green = int(red[2:4], 16) / RGBMAX
            red = int(red[0:2], 16) / RGBMAX
        else:
            red, green, blue = red
    h, l, s = rgb_to_hls(red, green, blue)
    return (int(round(h * HLSMAX)), int(round(l * HLSMAX)), int(round(s * HLSMAX)))


def ms_hls_to_rgb(hue, lightness=None, saturation=None):
    """Converts HLSMAX based HLS values to rgb values in the range (0,1)"""
    if lightness is None:
        hue, lightness, saturation = hue
    return hls_to_rgb(hue / HLSMAX, lightness / HLSMAX, saturation / HLSMAX)


def rgb_to_hex(red, green=None, blue=None):
    """Converts (0,1) based RGB values to a hex string 'rrggbb'"""
    if green is None:
        red, green, blue = red
    return (
        '%02x%02x%02x'
        % (int(round(red * RGBMAX)), int(round(green * RGBMAX)), int(round(blue * RGBMAX)))
    ).upper()


def get_theme_colors(wb):
    """Gets theme colors from the workbook"""
    # see: https://groups.google.com/forum/#!topic/openpyxl-users/I0k3TfqNLrc
    from openpyxl.xml.functions import QName, fromstring

    xlmns = 'http://schemas.openxmlformats.org/drawingml/2006/main'
    root = fromstring(wb.loaded_theme)
    themeEl = root.find(QName(xlmns, 'themeElements').text)
    colorSchemes = themeEl.findall(QName(xlmns, 'clrScheme').text)
    firstColorScheme = colorSchemes[0]

    colors = []
    for c in [
        'lt1',
        'dk1',
        'lt2',
        'dk2',
        'accent1',
        'accent2',
        'accent3',
        'accent4',
        'accent5',
        'accent6',
    ]:
        accent = firstColorScheme.find(QName(xlmns, c).text)
        if accent is not None and len(accent) > 0:
            child = accent[0]
            if 'window' in child.attrib['val']:
                colors.append(child.attrib['lastClr'])
            else:
                colors.append(child.attrib['val'])
    return colors


def tint_luminance(tint, lum):
    """Tints a HLSMAX based luminance"""
    # See: http://ciintelligence.blogspot.co.uk/2012/02/converting-excel-theme-color-and-tint.html
    if tint < 0:
        return int(round(lum * (1.0 + tint)))
    else:
        return int(round(lum * (1.0 - tint) + (HLSMAX - HLSMAX * (1.0 - tint))))


def theme_and_tint_to_rgb(wb, theme, tint):
    """Given a workbook, a theme number and a tint return a hex based rgb"""
    rgb = get_theme_colors(wb)[theme]
    h, l, s = rgb_to_ms_hls(rgb)
    return rgb_to_hex(ms_hls_to_rgb(h, tint_luminance(tint, l), s))


def convert_hex(hex_code):
    """
    Ensures that a hex color code is in a 6-digit format compatible with Matplotlib.
    If the hex code is 8 characters long, it assumes the first two characters are the alpha channel
    and removes them since Matplotlib will assume a default alpha of 1 (fully opaque) if not specified.

    :param hex_code: A hex color code (string)
    :return: A 6-digit hex color string compatible with Matplotlib
    """
    # Remove the alpha channel if the hex is in ARGB format
    if len(hex_code) == 8:
        hex_code = hex_code[2:]

    # Ensure the hex code starts with '#'
    if not hex_code.startswith('#'):
        hex_code = '#' + hex_code

    return hex_code


def get_cell_colors(workbook, sheet_name, min_row, min_col, max_row=None, max_col=None):
    """
    Extracts cell shading colors from a range of cells in an Excel worksheet.

    Parameters
    ----------
    workbook : openpyxl.Workbook
        The workbook containing the worksheet to extract colors from.
    sheet_name : str
        The name of the worksheet to extract colors from.
    min_row : int
        The first row to extract colors from.
    min_col : int
        The first column to extract colors from.
    max_row : int, optional
        The last row to extract colors from. If None, the last row in the worksheet is used.
    max_col : int, optional
        The last column to extract colors from. If None, the last column in the worksheet is used.

    Returns
    -------
    dict
        A dictionary with keys as (row, column) tuples and values as the color of the cell.

    """
    # Get the worksheet
    sheet = workbook[sheet_name]

    # If max_row or max_col is None, use the maximum row or column in the worksheet
    if max_row is None:
        max_row = sheet.max_row
    if max_col is None:
        max_col = sheet.max_column

    # Create an empty dictionary to store the colors
    c_dict = {}
    colors = []
    for row in sheet.iter_rows(min_row=min_row, max_row=max_row, min_col=min_col, max_col=max_col):
        for cell in row:
            color_idx = cell.fill.start_color.index

            if isinstance(color_idx, str):
                color = convert_hex(color_idx)
            elif isinstance(color_idx, int):
                theme = cell.fill.start_color.theme
                tint = cell.fill.start_color.tint
                color = theme_and_tint_to_rgb(workbook, theme, tint)
                color = convert_hex(color)
            else:
                color = 'white'
            # get column letter:
            col_letter = openpyxl.utils.get_column_letter(cell.column)
            # print(f'Cell {col_letter}{cell.row}: {color}')
            c_dict[(col_letter, cell.row)] = color
            colors.append(color)
    return colors, c_dict


def get_value_color_map(workbook, sheet_name, min_row, min_col, max_row=None, max_col=None):
    """
    Uses get_cell_colors to build a dictionary where each key is a cell's value
    and each value is the cell's color in hex format.

    Parameters
    ----------
    workbook : openpyxl.Workbook
        The workbook containing the worksheet to extract colors from.
    sheet_name : str
        The name of the worksheet to extract colors from.
    min_row : int
        The first row to extract colors from.
    min_col : int
        The first column to extract colors from.
    max_row : int, optional
        The last row to extract colors from. If None, the last row in the worksheet is used.
    max_col : int, optional
        The last column to extract colors from. If None, the last column in the worksheet is used.

    Returns
    -------
    dict
        A dictionary where each key is the cell value and each value is the color (hex code).
        If multiple cells share the same value, only the last color encountered is stored.
    """

    # Reuse the existing function to get the list of colors and dict of (column_letter, row) -> color
    colors, c_dict = get_cell_colors(workbook, sheet_name, min_row, min_col, max_row, max_col)

    # Retrieve the sheet so we can look up the actual cell values
    sheet = workbook[sheet_name]

    # Build a dictionary that maps cell values to color hex codes
    value_color_map = {}
    for (col_letter, row), color_hex in c_dict.items():
        col_index = column_index_from_string(col_letter)
        cell_value = sheet.cell(row=row, column=col_index).value

        # Store/overwrite the color keyed by the cell's value
        value_color_map[cell_value] = color_hex

    return value_color_map


def plot_colored_boxes(value_color_map, cols=4, box_size=(2, 1), figname=None):
    """
    Creates a visualization of the value_color_map dictionary.

    Parameters
    ----------
    value_color_map

    """

    keys = list(value_color_map.keys())
    rows = (len(keys) + cols - 1) // cols  # Compute required rows
    fig, ax = plt.subplots(figsize=(cols * box_size[0], rows * box_size[1]))
    ax.set_xlim(0, cols)
    ax.set_ylim(0, rows)
    ax.set_xticks([])
    ax.set_yticks([])
    ax.set_frame_on(False)

    for idx, key in enumerate(keys):
        row = rows - 1 - (idx // cols)  # Reverse row order for plotting
        col = idx % cols
        color = value_color_map[key]

        rect = plt.Rectangle((col, row), 1, 1, color=color, ec='black')
        ax.add_patch(rect)
        ax.text(
            col + 0.5,
            row + 0.5,
            key,
            ha='center',
            va='center',
            fontsize=12,
            weight='bold',
            color='white' if int(color[1:], 16) < 0x888888 else 'black',
        )

    if figname:
        fig.savefig(figname)

    return fig, ax
